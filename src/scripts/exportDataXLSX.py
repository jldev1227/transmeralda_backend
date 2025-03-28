#!/usr/bin/env python
import pandas as pd
import json
import sys
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

def flatten_json(nested_json, prefix=''):
    """
    Aplana un objeto JSON anidado.
    
    Args:
        nested_json (dict): Objeto JSON anidado
        prefix (str): Prefijo para las claves anidadas
        
    Returns:
        dict: Objeto JSON aplanado donde las claves anidadas se unen con '_'
    """
    flat_json = {}
    
    for key, value in nested_json.items():
        new_key = f"{prefix}_{key}" if prefix else key
        
        if isinstance(value, dict):
            # Recursivamente aplanar diccionarios anidados
            flat_json.update(flatten_json(value, new_key))
        elif isinstance(value, list):
            # Convertir listas a formato string JSON
            if value and isinstance(value[0], dict):
                # Si es una lista de diccionarios, intentar extraer información clave
                try:
                    # Extraer campos importantes si existen en los diccionarios
                    summary = []
                    for item in value:
                        if 'nombre' in item or 'name' in item:
                            name = item.get('nombre', item.get('name', ''))
                            value_field = item.get('valor', item.get('value', ''))
                            summary.append(f"{name}: {value_field}")
                        elif 'fecha' in item and 'valor' in item:
                            summary.append(f"{item['fecha']}: {item['valor']}")
                    
                    if summary:
                        flat_json[new_key] = "; ".join(summary)
                    else:
                        flat_json[new_key] = json.dumps(value, ensure_ascii=False)
                except:
                    # Si algo falla, simplemente convertir a JSON
                    flat_json[new_key] = json.dumps(value, ensure_ascii=False)
            else:
                # Para listas simples, unir elementos
                flat_json[new_key] = ", ".join(str(x) for x in value)
        else:
            # Valores simples
            flat_json[new_key] = value
            
    return flat_json

def custom_export_to_excel(json_array, output_path=None, company_name="Transmeralda"):
    """
    Exporta un array de liquidaciones a un archivo Excel con columnas personalizadas,
    agrupando por periodo_end y creando una hoja diferente para cada mes
    """
    # Si no hay datos, retornamos None
    if not json_array:
        print("No hay datos para exportar")
        return None
    
    print(f"Procesando {len(json_array)} liquidaciones...")
    
    # Agrupar las liquidaciones por periodo_end
    liquidaciones_por_periodo = {}
    
    # Mapeo de números de mes a nombres en español
    meses_espanol = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
        7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }
    
    for item in json_array:
        # Extraer y formatear la fecha de periodo_end
        periodo_end = item.get('periodo_end', '')
        
        print(item)
        
        if periodo_end:
            try:
                # Convertir la fecha a objeto datetime para extraer mes y año
                fecha_obj = datetime.strptime(periodo_end, "%Y-%m-%d")
                mes = fecha_obj.month
                anio = fecha_obj.year
                # Crear clave para agrupar (ej: "SEPTIEMBRE 2025")
                periodo_key = f"{meses_espanol[mes]} {anio}"
                
                print(periodo_key)
                
                # Añadir a la lista correspondiente
                if periodo_key not in liquidaciones_por_periodo:
                    liquidaciones_por_periodo[periodo_key] = []
                
                liquidaciones_por_periodo[periodo_key].append(item)
            except ValueError:
                # Si hay problemas con el formato de fecha, usar "SIN FECHA"
                if "SIN FECHA" not in liquidaciones_por_periodo:
                    liquidaciones_por_periodo["SIN FECHA"] = []
                liquidaciones_por_periodo["SIN FECHA"].append(item)
        else:
            # Para liquidaciones sin periodo_end
            if "SIN FECHA" not in liquidaciones_por_periodo:
                liquidaciones_por_periodo["SIN FECHA"] = []
            liquidaciones_por_periodo["SIN FECHA"].append(item)
    
    # Si no se especifica ruta, generamos una con timestamp
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"liquidaciones_nomina_{timestamp}.xlsx"
    
    print(f"Generando archivo Excel en: {output_path}")
    
    # Asegurar que el directorio existe
    dir_path = os.path.dirname(os.path.abspath(output_path))
    if dir_path:
        os.makedirs(dir_path, exist_ok=True)
    
    # Crear un nuevo libro de Excel
    wb = Workbook()
    
    # Eliminar la hoja predeterminada para empezar desde cero
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Crear una hoja para cada periodo
    for sheet_name, liquidaciones in liquidaciones_por_periodo.items():
        print(f"Creando hoja para período: {sheet_name}")
        
        # Crear una nueva hoja con el nombre del periodo
        ws = wb.create_sheet(title=sheet_name)
        
        # Preparar los datos para esta hoja
        data = []
        for idx, item in enumerate(liquidaciones):
            # Extraer datos del conductor
            conductor = item.get('conductor', {})
            nombre_completo = f"{conductor.get('nombre', '')} {conductor.get('apellido', '')}"
            
            # Calcular valores derivados
            salario_devengado = float(item.get('salario_devengado', 0))
            auxilio_transporte = float(item.get('auxilio_transporte', 0))
            salud = float(item.get('salud', 0))
            pension = float(item.get('pension', 0))
            total_anticipos = float(item.get('total_anticipos', 0))
            
            valor_a_liquidar = salario_devengado + auxilio_transporte
            total_deducciones = salud + pension
            total_a_pagar = valor_a_liquidar - total_deducciones
            total_a_pagar = valor_a_liquidar - total_deducciones
            
            novedad = item.get('observaciones', 'No especificada')
                
            # Verificar si el conductor es recién ingresado
            if conductor.get('fecha_ingreso'):
                try:
                    fecha_ingreso = datetime.strptime(conductor.get('fecha_ingreso'), '%Y-%m-%d')
                    fecha_inicio_liquidacion = datetime.strptime(item.get('periodo_start', '1900-01-01'), '%Y-%m-%d')
                    fecha_fin_liquidacion = datetime.strptime(item.get('periodo_end', '2999-12-31'), '%Y-%m-%d')
                    
                    # Verificar si la fecha de ingreso cae dentro del período de liquidación
                    if fecha_inicio_liquidacion <= fecha_ingreso <= fecha_fin_liquidacion:
                        novedad = "Recién ingresado"
                except (ValueError, TypeError) as e:
                    print(f"Error al procesar fecha de ingreso: {e}")
            
            # Verificar si el conductor tuvo vacaciones en este período
            if item.get('periodo_start_vacaciones') and item.get('periodo_end_vacaciones'):
                # Si ya tenía una novedad, añadimos "Vacaciones", de lo contrario, asignamos "Vacaciones"
                if novedad != 'No especificada' and novedad:
                    novedad += "; Vacaciones"
                else:
                    novedad = "Vacaciones"
                

            
            # Crear fila con los datos requeridos
            row = {
                'Indice': idx + 1,
                'Conductor': nombre_completo,
                'Identificación': conductor.get('numero_identificacion', ''),
                'Cargo': "Conductor",
                'Lugar de Trabajo': conductor.get('sede_trabajo', 'No especificado'),
                'Novedad': novedad,
                'Salario Base': conductor.get('salario_base', 0),
                'Fecha Ingreso': conductor.get('fecha_ingreso', ''),
                'Fecha Retiro': conductor.get('fecha_retiro', ''),
                'Días Laborados': item.get('dias_laborados', 0),
                'Salario Devengado': salario_devengado,
                'Auxilio Transporte': auxilio_transporte,
                'Valor a Liquidar': valor_a_liquidar,
                'Salud': salud,
                'Pensión': pension,
                'Total Deducciones': total_deducciones,
                'Anticipos': total_anticipos,
                'Total a Pagar Básico': total_a_pagar
            }
            data.append(row)
        
        # Crear DataFrame para esta hoja
        df = pd.DataFrame(data)
        
        # Configurar estilos
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="006B3C", end_color="006B3C", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(size=10)
        data_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        money_font = Font(size=10, bold=True)
        money_alignment = Alignment(horizontal='right', vertical='center')
        
        # Añadir encabezado con título y fecha
        ws.merge_cells('A1:R1')
        title_cell = ws['A1']
        title_cell.value = f"LIQUIDACIÓN DE NÓMINA OPERATIVA - {sheet_name}"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="006B3C", end_color="006B3C", fill_type="solid")
        
        # Añadir información de la empresa
        ws.merge_cells('A2:R2')
        company_cell = ws['A2']
        company_cell.value = company_name.upper()
        company_cell.font = Font(size=14, bold=True)
        company_cell.alignment = Alignment(horizontal='center', vertical='center')
        company_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        
        # Añadir fecha de generación
        ws.merge_cells('A3:R3')
        date_cell = ws['A3']
        date_cell.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        date_cell.font = Font(italic=True)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Dejar una fila en blanco
        current_row = 5
        
        # Añadir encabezados de columnas
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = header_fill
            cell.border = thin_border
        
        # Añadir los datos con formato
        for row_idx, row in enumerate(df.itertuples(index=False), 1):
            row_num = current_row + row_idx
            
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.value = value
                cell.border = thin_border
                
                # Formato para columnas monetarias
                if headers[col_idx-1] in ['Salario Base', 'Salario Devengado', 'Auxilio Transporte', 
                                        'Valor a Liquidar', 'Salud', 'Pensión', 'Total Deducciones', 
                                        'Anticipos', 'Total a Pagar Básico']:
                    cell.number_format = '"$"#,##0'
                    cell.font = money_font
                    cell.alignment = money_alignment
                    
                    # Colorear valores negativos en rojo
                    if isinstance(value, (int, float)) and value < 0:
                        cell.font = Font(size=10, bold=True, color="FF0000")
                else:
                    cell.font = data_font
                    cell.alignment = data_alignment
                
                # Colorear filas alternadas para mejorar la legibilidad
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Ajustar el ancho de las columnas
        for col_idx, column in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            if column in ['Conductor', 'Novedad']:
                ws.column_dimensions[col_letter].width = 40
            elif column in ['Lugar de Trabajo', 'Identificación']:
                ws.column_dimensions[col_letter].width = 20
            else:
                ws.column_dimensions[col_letter].width = 15
        
        # Aplicar filtro automático a los encabezados
        ws.auto_filter.ref = f"A{current_row}:{get_column_letter(len(headers))}{current_row + len(df)}"
        
        # Inmovilizar paneles para mantener los encabezados visibles al desplazarse
        ws.freeze_panes = f"A{current_row + 1}"
        
        # Añadir totales al final (solo si hay datos)
        if not df.empty:
            total_row = current_row + len(df) + 1
            
            # Merge para el texto "TOTALES"
            ws.merge_cells(f'A{total_row}:J{total_row}')
            total_label = ws[f'A{total_row}']
            total_label.value = "TOTALES"
            total_label.font = Font(size=11, bold=True)
            total_label.alignment = Alignment(horizontal='right', vertical='center')
            total_label.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            
            # Columnas que deben sumarse
            sum_columns = {
                'Salario Devengado': 'K',
                'Auxilio Transporte': 'L',
                'Valor a Liquidar': 'M',
                'Salud': 'N',
                'Pensión': 'O',
                'Total Deducciones': 'P',
                'Anticipos': 'Q',
                'Total a Pagar': 'R'
            }
            
            # Añadir las sumas con formato
            for col_name, col_letter in sum_columns.items():
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1
                    cell = ws.cell(row=total_row, column=col_idx)
                    start_row = current_row + 1
                    end_row = total_row - 1
                    cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{end_row})"
                    cell.font = Font(size=11, bold=True)
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                    cell.number_format = '"$"#,##0'
                    cell.border = thin_border
            
            # Añadir pie de página
            footer_row = total_row + 2
            ws.merge_cells(f'A{footer_row}:R{footer_row}')
            footer_cell = ws[f'A{footer_row}']
            footer_cell.value = "Documento generado automáticamente - Sistema de Gestión TRANSMERALDA"
            footer_cell.font = Font(italic=True, size=8)
            footer_cell.alignment = Alignment(horizontal='center')
    
    # Ordenar las pestañas alfabéticamente
    wb._sheets.sort(key=lambda x: x.title)
    
    # Guardar el archivo
    wb.save(output_path)
    print(f"Archivo exportado exitosamente a: {output_path}")
    
    return output_path

def main():
    """Función principal que ejecuta el script desde línea de comandos"""
    try:
        if len(sys.argv) < 2:
            print("Uso: python liquidaciones_export.py <tempFilePath> [output_path]")
            sys.exit(1)
        
        # Obtener la ruta del archivo JSON
        temp_file_path = sys.argv[1]
        print(f"Leyendo archivo JSON: {temp_file_path}")
        
        # Leer el contenido del archivo
        with open(temp_file_path, 'r', encoding='utf-8') as file:
            json_string = file.read()
            
        # Cargar el JSON desde el contenido del archivo
        json_array = json.loads(json_string)
        print(f"JSON cargado correctamente, tipo: {type(json_array)}")
        
        # Validar que sea un array
        if not isinstance(json_array, list):
            print("Convertido objeto único a array")
            # Si no es un array, lo convertimos en uno
            json_array = [json_array]
        
        # Ruta opcional de salida (argumento 2)
        output_path = sys.argv[2] if len(sys.argv) > 2 else None
        
        # Exportar datos con formato personalizado
        result_path = custom_export_to_excel(json_array, output_path)
        
        if result_path:
            # Devolvemos la ruta en la salida estándar para que el controlador pueda capturarla
            print(result_path)
            sys.exit(0)
        else:
            print("No se generó archivo de salida")
            sys.exit(1)
            
    except json.JSONDecodeError as e:
        print(f"Error al decodificar JSON: {str(e)}")
        sys.exit(1)
    except Exception as e:
        print(f"Error inesperado: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
if __name__ == "__main__":
    main()